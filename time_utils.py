# *_*coding:utf-8 *_*
# @Author : Reggie
# @Time : 2022/10/11 14:00
import datetime
import json
import logging
import os
import time
from functools import wraps
from pathlib import Path

# from io import StringIO

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
# print(BASE_DIR)
logger_path = Path(BASE_DIR) \
    .joinpath("logs") \
    .joinpath("time_cost_statistic.log")
time_cost_statistic_logger = logging.getLogger('time_cost_statistic')
formatter = logging.Formatter('%(message)s')
handler = logging.FileHandler(logger_path.as_posix(), encoding='utf-8')
handler.setFormatter(formatter)
time_cost_statistic_logger.addHandler(handler)
time_cost_statistic_logger.setLevel("INFO")


def reset_value_type(value):
    if isinstance(value, (list, tuple, set)):
        result = []
        for v in value:
            result.append(reset_value_type(v))
        return result
    elif isinstance(value, dict):
        result = {}
        for k, v in value.items():
            result[reset_value_type(k)] = reset_value_type(v)
        return result
    elif isinstance(value, str):
        return value
    else:
        return str(value)


def time_cost_statistic(func):
    @wraps(func)
    def inner(*args, **kwargs):
        time_start = time.time()
        # s_io = StringIO()
        # dis.dis(func, file=s_io)
        # s_io.seek(0)
        # dis_info = s_io.read()
        # s_io.close()
        out = func(*args, **kwargs)
        time_stop = time.time()
        info = {
            "start_time": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time_start)),
            "stop_time": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time_stop)),
            "start_timestamp": time_start,
            "stop_timestamp": time_stop,
            "cost_timestamp": time_stop - time_start,
            "func_name": func.__name__,
            "type": str(func),
            "args": [reset_value_type(i) for i in args],
            "kwargs": {reset_value_type(k): str(v) for k, v in kwargs.items()},
            "out": reset_value_type(out),
            # "dis_info": dis_info,
        }
        time_cost_statistic_logger.info(f"{json.dumps(info, ensure_ascii=False)}")
        return out

    return inner


class BaseStatistic:
    def __init__(self, *args, **kwargs):
        self.cost_statistics = self.get_all_cost_statistics()

    @classmethod
    def get_all_cost_statistics(cls):
        result = []
        with open(logger_path.as_posix(), encoding="utf-8") as f:
            for line in f.readlines():
                data = json.loads(line.strip())
                result.append(data)
        return result

    @classmethod
    def search_gt_time_cost_statistic(cls, cost_statistics, max_time_cost):
        result = []
        for cost_statistic in cost_statistics:
            if cost_statistic['cost_timestamp'] >= max_time_cost:
                result.append(cost_statistic)
            print(cost_statistic)
        return result

    @classmethod
    def sort_time_cost_statistic(cls, cost_statistics, reverse=False):
        return sorted(cost_statistics, key=lambda x: x['cost_timestamp'], reverse=reverse)

    @classmethod
    def filter_numb_for_cost_statistics(cls, cost_statistics, numb, reverse=False):
        return cls.sort_time_cost_statistic(cost_statistics, reverse=reverse)[:numb]

    def sort(self, reverse=False):
        self.cost_statistics = self.sort_time_cost_statistic(self.cost_statistics, reverse=reverse)


class CostStatistic(BaseStatistic):
    def __init__(self):
        super().__init__()

    def show(self):
        cost_statistics = self.cost_statistics
        keys = [
            "start_time",
            "cost_timestamp",
            "func_name",
            "type",
            "args",
            "kwargs",
            "out",
        ]
        max_length = 40
        max_key_dict = {}
        for cost_statistic in cost_statistics:
            for k in keys:
                max_key_dict.setdefault(k, len(k))
                value_length = len(str(cost_statistic[k]).strip())
                if max_key_dict[k] < value_length:
                    max_key_dict[k] = min(len(str(cost_statistic[k]).strip()), max_length)

        max_index = max(len(str(len(self.cost_statistics))) + 2, len("index") + 2)
        s = f"|{'index'.center(max_index, ' ')}|"
        for k in keys:
            s += k.center(max_key_dict[k] + 2, " ") + " |"
        print(s)

        for index, cost_statistic in enumerate(cost_statistics):
            s = f"|{str(index).center(max_index, ' ')}| "
            # s = "| "
            for k in keys:
                v = str(cost_statistic[k]).strip()[:max_length]
                s += v.center(max_key_dict[k] + 2, " ") + " |"
            print(s)


class ExportExcelStatistic(BaseStatistic):

    def __init__(self, dist_name=None, dist_path=None):
        """

        Args:
            dist_name: 输出文件名称
            dist_path: 输出文件路径
        """
        self.dist_name = dist_name or f"logs/Cost statistics{datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S')}.xlsx"
        self.dist_path = dist_path or BASE_DIR
        self.dist = Path(self.dist_path).joinpath(self.dist_name)
        self.dist.parent.mkdir(exist_ok=True, parents=True)
        super().__init__()

    def save(self):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, Side, Border

        wb = openpyxl.Workbook(self.dist)
        wb.save(self.dist)
        wb = openpyxl.load_workbook(self.dist)
        sheet = wb.active
        sheet.title = "Cost statistics"
        keys = [
            "start_time",
            "cost_timestamp",
            "func_name",
            "type",
            "args",
            "kwargs",
            "out",
        ]
        max_index = max(len(str(len(self.cost_statistics))) + 2, len("index") + 2)
        for i, k in enumerate(keys):
            sheet.cell(1, i + 2).value = k

        max_key_dict = {}
        max_value_length = 100
        for index, cost_statistic in enumerate(self.cost_statistics):
            # 写入索引
            sheet.cell(index + 2, 1).value = index + 1
            for i, k in enumerate(keys):
                # 写入表格
                value = cost_statistic[k]
                if isinstance(value, (int, float)):
                    value = value
                else:
                    value = str(value)
                sheet.cell(index + 2, i + 2).value = value
                # 统计长度
                max_key_dict.setdefault(k, len(k))
                value_length = len(str(cost_statistic[k]).strip())
                if max_key_dict[k] < value_length:
                    if k == "out" and value_length > max_value_length:
                        value_length = min(value_length, max_value_length)
                    max_key_dict[k] = value_length

        # 设置索引列宽
        col_letter = get_column_letter(1)
        sheet.column_dimensions[col_letter].width = max_index + 3
        # 设置列宽
        for i, k in enumerate(keys):
            # 根据列的数字返回字母
            col_letter = get_column_letter(i + 2)
            sheet.column_dimensions[col_letter].width = max_key_dict[k] + 3

        # 设置行高
        for i in range(1, sheet.max_row + 1):
            sheet.row_dimensions[i].height = 14.4

        fill = openpyxl.styles.PatternFill("solid", fgColor="D9D9D9")  # 颜色也可以直接设置red等
        fill_read = openpyxl.styles.PatternFill("solid", fgColor="00FF0000")  # 颜色也可以直接设置red等
        thin_side = Side(style="thin")
        thin_border = Border(left=thin_side, bottom=thin_side, right=thin_side, top=thin_side)
        # 设置单元格对齐格式
        for i, col in enumerate(sheet.columns):
            for ii, cell in enumerate(col):
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if ii == 0:
                    cell.font = Font(size=10, bold=True)
                    cell.fill = fill
                    cell.border = thin_border
                else:
                    cell.font = Font(size=10)
        # 设置行样式
        for i, row in enumerate(sheet.rows):
            for ii, cell in enumerate(row):
                if ii == 0:
                    # 设置索引行样式
                    cell.font = Font(size=10, bold=True)
                    cell.border = thin_border
                    cell.fill = fill
                if ii == 3 and "Run" in cell.value:
                    for iii, cell1 in enumerate(row):
                        if iii == 0:
                            continue
                        cell1.fill = fill_read

        wb.save(self.dist)

    def save2(self):
        import pandas as pd
        df = pd.DataFrame(self.cost_statistics)
        df.to_excel("test.xlsx")
        reset_col("test.xlsx")


def reset_col(filename):
    import pandas as pd
    import openpyxl
    from openpyxl.utils import get_column_letter
    wb = openpyxl.load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        df = pd.read_excel(filename, sheet).fillna('-')
        df.loc[len(df)] = list(df.columns)
        for col in df.columns:
            index = list(df.columns).index(col)
            letter = get_column_letter(index + 1)
            collen = df[col].apply(lambda x: len(str(x).encode())).max()
            ws.column_dimensions[letter].width = collen * 1.2 + 4

    wb.save(filename)


@time_cost_statistic
def t(a=1):
    time.sleep(2)
    return f"success:{a}"


if __name__ == '__main__':
    # from random import random, randint
    #
    # for i in range(0, 10):
    #     t(float(f"{randint(1, 2) + random():.2f}"))

    cost_statistics = CostStatistic()
    # cost_statistics.sort()
    cost_statistics.show()

    export_excel_statistics = ExportExcelStatistic()
    # export_excel_statistics.sort()
    export_excel_statistics.save()
